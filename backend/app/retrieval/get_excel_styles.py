from openpyxl import load_workbook
import logging
import sys
# Load workbook and select sheet
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from copy import copy


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    stream=sys.stdout
)

def copy_style_from_template_row(template_row, target_row):
    """Copy styles from a template row to a target row."""
    for template_cell, target_cell in zip(template_row, target_row):
        if isinstance(template_cell, Cell) and template_cell.has_style:
            target_cell.font = copy(template_cell.font)
            target_cell.fill = copy(template_cell.fill)
            target_cell.border = copy(template_cell.border)
            target_cell.alignment = copy(template_cell.alignment)
            target_cell.number_format = copy(template_cell.number_format)

def apply_styles_to_output(template_path, output_path, header_col_index=1, start_row=13):
    template_wb = load_workbook(template_path)
    output_wb = load_workbook(output_path)

    template_ws = template_wb.active
    output_ws = output_wb.active

    # Build a mapping from header value to its styled row in the template
    header_style_map = {}
    for row in template_ws.iter_rows(min_row=13):  # Skip header row in template
        cell = row[header_col_index - 1]
        if cell.value:
            header_key = str(cell.value).strip()
            if header_key not in header_style_map:
                header_style_map[header_key] = row

    # Apply styles to matching rows in output file
    for row in output_ws.iter_rows(min_row=start_row):
        cell = row[header_col_index - 1]
        if cell.value:
            header_key = str(cell.value).strip()
            if header_key in header_style_map:
                copy_style_from_template_row(header_style_map[header_key], row)

    # Save the updated workbook (you can change filename to avoid overwrite)
    output_wb.save(output_path)
 
def get_excel_styles(input_file_path, output_file_path):
    apply_styles_to_output(input_file_path, output_file_path, header_col_index=1, start_row=13)
# def get_excel_styles(input_file_path):
#     wb = load_workbook("shared_data/input_templates/Set II_MD_Irrigation Set.xlsx")
#     sheet = wb.active  # Or wb["SheetName"]

#     # Choose the row you want to inspect (e.g., row 13)
#     row_number = 13

#     # Get style info from each cell in the row
#     for col in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row=row_number, column=col)

#         logging.info(f"Cell {cell.coordinate}:")
#         logging.info(f"Font: {cell.font}")
#         logging.info(f"Fill: {cell.fill}")
#         logging.info(f"Border: {cell.border}")
#         logging.info(f"Alignment: {cell.alignment}")
#         logging.info(f" Number Format: {cell.number_format}")
#         logging.info(f"Protection: {cell.protection}")