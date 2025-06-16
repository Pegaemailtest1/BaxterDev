import openpyxl
from retrieval.questions.question_loader import questions, priority
from .process_question_templates import process_question_templates
from .retrieve_content_template import retrieve_collection_data_template
import json
import logging
import sys
from langchain.globals import set_llm_cache
from langchain_community.cache import InMemoryCache
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from copy import copy
import ast


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    stream=sys.stdout
)

dict_data = defaultdict(list)

def query_llama(question, document_id, prompt_template, full_document_search, where_filter, OLLAMA_URL, EMBED_MODEL,CHROMA_HOST,CHROMA_PORT,COLLECTION_NAME,FM_MODEL,MAX_RESULTS,temperature, max_tokens):
    
    llama_response =retrieve_collection_data_template(
        CHROMA_HOST,
        CHROMA_PORT,
        COLLECTION_NAME,
        EMBED_MODEL,
        question,
        document_id,
        prompt_template,
        full_document_search,
        where_filter,
        MAX_RESULTS,
        FM_MODEL,
        OLLAMA_URL,
        temperature,
        max_tokens
    )
    return extract_response_array(llama_response)


def extract_response_array(llama_response):
    try:
        # Clean and isolate JSON
        json_start = llama_response.index('{')
        json_data = llama_response[json_start:]
        logging.info(f"json data before: {json_data}")
        # Parse JSON
        data = json.loads(json_data)
        logging.info(f"json data: {data}")

        # Extract and return just the list
        response_list = data.get("response", "")
        logging.info(f"response_list: {response_list}")
        # Make sure it's actually a list
        if isinstance(response_list, list):
            logging.info(f"response data: {response_list}")
            return response_list
        else:
            logging.warning("response is not a list")
            return str(response_list)
    except (ValueError, json.JSONDecodeError, KeyError) as e:
        logging.warning(f"Failed to extract array: {e}")
        return llama_response.strip()  # fallback to list with raw string


def extract_product_code(sheet):
    """Find the cell that contains 'product code' and return the adjacent cell value"""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "product code" in cell.value.lower():
                adjacent_value = sheet.cell(row=cell.row, column=cell.column + 1).value
                return str(adjacent_value).strip() if adjacent_value else None
    return None

def write_list_to_repeating_rows(sheet, data, start_row, column): 
    for i, item in enumerate(data):
        sheet.cell(row=start_row + i, column=column).value = item

def copy_row_style(template_row, target_row):
    for template_cell, target_cell in zip(template_row, target_row):
        if isinstance(template_cell, Cell) and template_cell.has_style:
            target_cell.font = copy(template_cell.font)
            target_cell.fill = copy(template_cell.fill)
            target_cell.border = copy(template_cell.border)
            target_cell.alignment = copy(template_cell.alignment)
            target_cell.number_format = copy(template_cell.number_format)

    # Copy row height
    target_row_idx = target_row[0].row
    template_row_idx = template_row[0].row
    target_ws = target_row[0].parent
    template_ws = template_row[0].parent

    target_ws.row_dimensions[target_row_idx].height = template_ws.row_dimensions[template_row_idx].height

def apply_styles_by_first_column_match(input_template_path, output_file_path, start_row):
    # Load workbooks
    template_wb = openpyxl.load_workbook(input_template_path)
    output_wb = openpyxl.load_workbook(output_file_path)

    template_ws = template_wb.active
    output_ws = output_wb.active

    # Build map from first cell values in template (starting at row 13)
    template_style_map = {}
    for row in template_ws.iter_rows(min_row=start_row):
        key_cell = row[0]  # Column A
        if key_cell.value:
            key = str(key_cell.value).strip()
            template_style_map[key] = row

    # Apply styles to matching rows in output (starting at row 13)
    for row in output_ws.iter_rows(min_row=start_row):
        key_cell = row[0]  # Column A
        if key_cell.value:
            key = str(key_cell.value).strip()
            if key in template_style_map:
                copy_row_style(template_style_map[key], row)

    output_wb.save(output_file_path)


def form_dict_input(input_type, input_value, column, row, repeat, q_type):
    if q_type == "group" and isinstance(input_value, list):
        input_value = ", ".join(input_value)

    if isinstance(input_value, list):
        for item in input_value:
            if item.strip().lower() == "no trace":
                continue
            found = False
            for entry in dict_data[input_type]:
                if entry["value"] == item:
                    entry["column"].append(column)
                    found = True
                    break
            if not found:
                dict_data[input_type].append({
                    "value": item,
                    "column": [column]
                })

    elif isinstance(input_value, str):
        new_entry = {
            "value": input_value,
            "column": column
        }
        if row is not None:
            new_entry["row"] = row
        if repeat:
            new_entry["repeat"] = repeat
        dict_data[input_type].append(new_entry)

    else:
        print(f"Unsupported data type for {input_type}: {type(input_value)}")


def generate_set2_template(input_file_path, output_file_path, OLLAMA_URL, EMBED_MODEL,CHROMA_HOST,CHROMA_PORT,COLLECTION_NAME,FM_MODEL,MAX_RESULTS,temperature, max_tokens):
    # Set InMemoryCache as the global LLM cache
    set_llm_cache(InMemoryCache())

    # logging.info(f"generate output template")
    # Load workbook and sheet    
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active

    priority_list = priority()
    # Step 1: Extract product code
    product_code = extract_product_code(sheet)
    if not product_code:
        return
            
    # Step 2: Generate question templates
    process_question_templates(
        questions_fn=questions,                      
        product_code=product_code,                   
        form_dict_input_fn=form_dict_input,          
        query_llama_fn=query_llama,                 
        dict_data=dict_data,                        
        config={                                    
            "OLLAMA_URL": OLLAMA_URL,
            "EMBED_MODEL": EMBED_MODEL,
            "CHROMA_HOST": CHROMA_HOST,
            "CHROMA_PORT": CHROMA_PORT,
            "COLLECTION_NAME": COLLECTION_NAME,
            "FM_MODEL": FM_MODEL,
            "MAX_RESULTS": MAX_RESULTS,
            "temperature": temperature,
            "max_tokens": max_tokens
        }
    )

    #logging.info(f"priority:{priority_list}")

    for item in priority_list:
        entries = dict_data.get(item, [])

        for i, entry in enumerate(entries):
            column = entry.get("column")
            value = entry.get("value")
            row = entry.get("row", None)
            repeat = entry.get("repeat", None)
            # Case 1: Specific row and column are defined
            if row is not None and column is not None:
                if isinstance(column, list):
                    for col in column:
                        sheet.cell(row=row, column=col).value = value
                else:
                    sheet.cell(row=row, column=column).value = value
                sheet.cell(row=row, column=1).value = item  # optional: set item as header

            # Case 2: Repeat value for each row in range
            elif repeat == "True" and column is not None:
                if isinstance(column, list):
                    for row_idx in range(13, sheet.max_row + 1):
                        for col in column:
                            sheet.cell(row=row_idx, column=col).value = value
                            sheet.cell(row=row_idx, column=1).value = item
                else:
                    for row_idx in range(13, sheet.max_row + 1):
                        sheet.cell(row=row_idx, column=column).value = value

            # Case 3: Insert into next available row
            elif isinstance(column, list):
                next_row = get_next_empty_row(sheet, start=13)
                for col in column:
                    sheet.cell(row=next_row, column=col).value = value
                sheet.cell(row=next_row, column=1).value = item  # insert row label

            # Case 4: Unknown configuration - log and skip
            else:
                logging.info(f"Skipping entry at index {i}: row={row}, column={column}, value={value}, repeat={repeat}")

            
    
    for row in sheet.iter_rows(min_row=13, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if cell.column == 2:
                cell.value = product_code
            elif cell.value is None or str(cell.value).strip() == "":
                cell.value = "No Trace"
    
    # Save updated workbook
    workbook.save(output_file_path)
    apply_styles_by_first_column_match(input_file_path, output_file_path, start_row=13)
    logging.info(f"Done! Output saved to: {output_file_path}")

def get_next_empty_row(sheet, start):
    """Finds the first completely empty row starting from a given row."""
    row = start
    while True:
        if all(sheet.cell(row=row, column=col).value is None for col in range(3, sheet.max_column + 1)):
            return row
        row += 1