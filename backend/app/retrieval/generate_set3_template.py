import openpyxl
import logging
import sys
import ast
from collections import defaultdict
from openpyxl.utils import column_index_from_string
from langchain.globals import set_llm_cache
from langchain_community.cache import InMemoryCache
from .questions_set3 import questions
from .retrieve_content_template import retrieve_collection_data_template
import re
import json

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    stream=sys.stdout
)

dict_data = defaultdict(list)

def extract_product_code(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "product code" in cell.value.lower():
                adjacent_value = sheet.cell(row=cell.row, column=cell.column + 1).value
                return str(adjacent_value).strip() if adjacent_value else None
    return None

def extract_response_array(llama_response):
    try:
        # Extract content between triple backticks, fallback to entire string
        code_block = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", llama_response, re.DOTALL)
        json_data = code_block.group(1) if code_block else llama_response.strip()

        # Attempt to load JSON
        return json.loads(json_data)
    except Exception as e:
        logging.warning(f"Failed to extract array: {e}")
        return []

def query_llama(question, document_id, prompt_template, full_document_search, where_filter,
                OLLAMA_URL, EMBED_MODEL, CHROMA_HOST, CHROMA_PORT, COLLECTION_NAME,
                FM_MODEL, MAX_RESULTS, temperature, max_tokens):
    llama_response = retrieve_collection_data_template(
        CHROMA_HOST, CHROMA_PORT, COLLECTION_NAME, EMBED_MODEL,
        question, document_id, prompt_template, full_document_search, where_filter,
        MAX_RESULTS, FM_MODEL, OLLAMA_URL, temperature, max_tokens
    )
    return extract_response_array(llama_response)

def merge_iso_results_by_description(*document_results):
    merged = defaultdict(dict)
    description_order = []

    for idx, doc_data in enumerate(document_results):
        code_key = f"Code_{idx+1}"
        for item in doc_data:
            desc = item["description"].strip()
            code = item["code"].strip()
            if desc not in merged:
                description_order.append(desc)
            merged[desc][code_key] = code

    all_code_keys = [f"Code_{i+1}" for i in range(len(document_results))]
    final_output = []
    for desc in description_order:
        row = {"description": desc}
        for key in all_code_keys:
            row[key] = merged[desc].get(key, "N/A")
        final_output.append(row)

    return final_output

def generate_set3_template(input_file_path, output_file_path,
                           OLLAMA_URL, EMBED_MODEL, CHROMA_HOST, CHROMA_PORT, COLLECTION_NAME,
                           FM_MODEL, MAX_RESULTS, temperature, max_tokens):

    set_llm_cache(InMemoryCache())
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active

    product_code = extract_product_code(sheet)
    if not product_code:
        logging.warning("No product code found.")
        return

    document_outputs = []

    try:
        question_templates = questions()
        for q_template in question_templates["Questions"]:
            question_text = q_template["question"].format(product_code=product_code)
            document_id = q_template["document_id"]
            prompt_template = q_template["prompt_template"]
            full_document_search = q_template["full_document_search"]
            where_filter = q_template["where_filter"]
            q_max_results = q_template.get("max_results")
            max_results_override = int(q_max_results) if q_max_results else MAX_RESULTS

            response = query_llama(
                question_text, document_id, prompt_template, full_document_search,
                where_filter, OLLAMA_URL, EMBED_MODEL, CHROMA_HOST, CHROMA_PORT,
                COLLECTION_NAME, FM_MODEL, max_results_override, temperature, max_tokens
            )

            logging.info(f"Response for question '{question_text}': {response}")
            if isinstance(response, list):
                document_outputs.append(response)

    except Exception as e:
        logging.error(f"Error during LLaMA ISO query: {e}")

    logging.info(f"Document outputs: {document_outputs}")
    merged_data = merge_iso_results_by_description(*document_outputs)
    logging.info(f"Merged data: {merged_data}")

    output_start_row = 12
    for i, row_data in enumerate(merged_data):
        sheet.cell(row=output_start_row + i, column=2).value = row_data["description"]
        sheet.cell(row=output_start_row + i, column=3).value = row_data.get("Code_1", "N/A")
        sheet.cell(row=output_start_row + i, column=4).value = row_data.get("Code_2", "N/A")
        sheet.cell(row=output_start_row + i, column=5).value = row_data.get("Code_3", "N/A")
        sheet.cell(row=output_start_row + i, column=6).value = row_data.get("Code_4", "N/A")

    workbook.save(output_file_path)
    logging.info(f"Done! Output saved to: {output_file_path}")
