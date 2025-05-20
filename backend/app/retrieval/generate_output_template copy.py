import openpyxl
from .questions import questions, priority # Your function that returns question templates
from .retrieve_content import retrieve_collection_data
import json
import logging
import sys
from langchain.globals import set_llm_cache
from langchain.cache import InMemoryCache
from collections import defaultdict
from openpyxl.utils import get_column_letter
from copy import copy
import ast
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    stream=sys.stdout
)

dict_data = defaultdict(list)

def query_llama(question, OLLAMA_URL, EMBED_MODEL,CHROMA_HOST,CHROMA_PORT,COLLECTION_NAME,FM_MODEL,MAX_RESULTS,temperature, max_tokens):
    
    llama_response =retrieve_collection_data(
        CHROMA_HOST,
        CHROMA_PORT,
        COLLECTION_NAME,
        EMBED_MODEL,
        question,
        MAX_RESULTS,
        FM_MODEL,
        OLLAMA_URL,
        temperature,
        max_tokens
    )
    return extract_response_array(llama_response)


def extract_response_array(llama_response):
    try:
        # Clean and isolate JSON
        json_start = llama_response.index('{')
        json_data = llama_response[json_start:]
        logging.info(f"json data before: {json_data}")
        # Fix common LLM formatting issues
        # json_data = json_data.replace('""', '"')   # remove double quotes
        # json_data = json_data.replace('\\"', '"')
        # json_data = json_data.replace("\n", "")
        # json_data = json_data.replace('\\n\\n', "")
        # json_data = json_data.replace('\\"\\"', '"')
        # json_data = json_data.replace('\n\"', '"')
        # json_data = json_data.strip()
        
        # Parse JSON
        data = json.loads(json_data)
        logging.info(f"json data: {data}")

        # Extract and return just the list
        response_list = data.get("response", "")
        logging.info(f"response_list: {response_list}")
        # Make sure it's actually a list
        if isinstance(response_list, list):
            logging.info(f"response data: {response_list}")
            return response_list
        else:
            logging.warning("response is not a list")
            return str(response_list)
    except (ValueError, json.JSONDecodeError, KeyError) as e:
        logging.warning(f"Failed to extract array: {e}")
        return llama_response.strip()  # fallback to list with raw string


def extract_product_code(sheet):
    """Find the cell that contains 'product code' and return the adjacent cell value"""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "product code" in cell.value.lower():
                adjacent_value = sheet.cell(row=cell.row, column=cell.column + 1).value
                return str(adjacent_value).strip() if adjacent_value else None
    return None

def write_list_to_repeating_rows(sheet, data, start_row, column): 
    for i, item in enumerate(data):
        sheet.cell(row=start_row + i, column=column).value = item

def insert_row_with_styles(ws, target_row_index):
    # Insert a blank row first
    ws.insert_rows(target_row_index)

    # Get the row to copy styles from (the one above the inserted row)
    template_row = ws[target_row_index - 1]

    for col_idx, cell in enumerate(template_row, start=1):
        new_cell = ws.cell(row=target_row_index, column=col_idx)
        new_cell.value = ""  # Optional: you can set your own value here
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

def form_dict_input(input_type, input_value, column, row, repeat):
    if isinstance(input_value, list):
        for item in input_value:
            found = False
            for entry in dict_data[input_type]:
                if entry["value"] == item:
                    entry["column"].append(column)
                    found = True
                    break
            if not found:
                dict_data[input_type].append({
                    "value": item,
                    "column": [column]
                })
    elif isinstance(input_value, str):
        # This block is for descriptive fields like Intended Use, etc.
        new_entry = {
            "value": input_value,
            "column": column
        }
        if row is not None:
            new_entry["row"] = row
           
        if repeat is not None:
            new_entry["repeat"] = repeat

        dict_data[input_type].append(new_entry)

    else:
        print(f"Unsupported data type for {input_type}: {type(input_value)}")

def generate_output_template(input_file_path, output_file_path, OLLAMA_URL, EMBED_MODEL,CHROMA_HOST,CHROMA_PORT,COLLECTION_NAME,FM_MODEL,MAX_RESULTS,temperature, max_tokens):
    # Set InMemoryCache as the global LLM cache
    set_llm_cache(InMemoryCache())

    # logging.info(f"generate output template")
    # Load workbook and sheet    
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active

    priority_list = priority()
    # Step 1: Extract product code
    product_code = extract_product_code(sheet)
    if not product_code:
        return
            
    # Step 2: Generate question templates
    try:
        question_templates = questions()
        # logging.info(f"question templates:{question_templates}")
        # Step 2d: Generate and write answers in columns C onward
        for i, q_template in enumerate(question_templates["Questions"]):
            question_text = q_template["question"].format(product_code=product_code)
            # logging.info(f"question text:{question_text}")

            column = int(q_template["column"])
            # logging.info(f"column:{column}")
            
            row = int(q_template["row"]) if "row" in q_template else None
            # logging.info(f"row:{row}")

            repeat = q_template["repeat"] if "repeat" in q_template else False
            dictionary_element = q_template["dictionary_element"]

            response = query_llama(question_text, OLLAMA_URL, EMBED_MODEL,CHROMA_HOST,CHROMA_PORT,COLLECTION_NAME,FM_MODEL,MAX_RESULTS,temperature, max_tokens)
            #response = "0719004306_IFU, Not found, 04-DEC-2024"
            #logging.info(f"llama response: {response}")
            
            if not response:
                continue

            if isinstance(response, str):
                try:
                    parsed = ast.literal_eval(response.strip())
                    response = parsed
                except Exception:
                    pass  # keep as original string if not parsable

            # Case 2: Normalize
            if isinstance(response, list):
                response = response
            elif isinstance(response, str):
                response = response.strip()
            else:
                response = ""
            try:
            # Now check type and call form_dict_input accordingly
                form_dict_input(dictionary_element, response, column, row, repeat)
                
            except Exception as e:
                logging.info(f"Error at dictionary element {dictionary_element}, question text {question_text}: {e}")
                continue

    except Exception as e:
        logging.info(f"Error to process input template: {e}")

            
    logging.info(f"dict data:{json.dumps(dict_data, indent=2)}")
    
    
    #logging.info(f"priority:{priority_list}")

    for item in priority_list:
        entries = dict_data.get(item, [])
        for i, entry in enumerate(entries):
            column = entry["column"]
            value = entry["value"]
            
            if "row" in entry and entry["row"] is not None:
                row = entry["row"]

            if "repeat" in entry and entry["repeat"] is not None:
                repeat = entry["repeat"]
            
            if isinstance(column, list) and row is None:
                next_row = get_next_empty_row(sheet, start=13)
                for col in column:
                    sheet.cell(row=next_row, column=col).value = value
                    sheet.cell(row=next_row, column=1).value = item

            if repeat == "True" and column is not None:
                for row in range(13, sheet.max_row + 1):
                    sheet.cell(row=row, column=column).value = value
            
            elif row is not None and column is not None:
                sheet.cell(row=row, column=column).value = value

            else:
                logging.info(i)
            
    
    for row in sheet.iter_rows(min_row=13, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            if cell.column == 2:
                # Forcefully set column 2 value as product code
                cell.value = product_code
            elif cell.value is None or str(cell.value).strip() == "":
                # Fill empty cells with "No Trace"
                cell.value = "No Trace"
    
    # Save updated workbook
    workbook.save(output_file_path)

    apply_styles_to_output(
        input_file_path,
        output_file_path,
        header_col_index=1,  # Column A
        start_row=16         # Start applying from row 16
    )
    logging.info(f"Done! Output saved to: {output_file_path}")


def get_next_empty_row(sheet, start):
    """Finds the first completely empty row starting from a given row."""
    row = start
    while True:
        if all(sheet.cell(row=row, column=col).value is None for col in range(3, sheet.max_column + 1)):
            return row
        row += 1


def copy_style_from_template_row(template_row, target_row):
    """Copy styles from a template row to a target row."""
    for template_cell, target_cell in zip(template_row, target_row):
        if template_cell.has_style:
            target_cell.font = template_cell.font
            target_cell.fill = template_cell.fill
            target_cell.border = template_cell.border
            target_cell.alignment = template_cell.alignment

def apply_styles_to_output(template_path, output_path, header_col_index, start_row):
    template_wb = load_workbook(template_path)
    output_wb = load_workbook(output_path)

    template_ws = template_wb.active
    output_ws = output_wb.active

    # Build a mapping from header value to its styled row in the template
    header_style_map = {}
    for row in template_ws.iter_rows(min_row=2):  # Skip template file header
        header_value = row[header_col_index - 1].value
        if header_value:
            header_style_map[str(header_value).strip()] = row

    # Apply styles to output rows starting from row 13
    for row in output_ws.iter_rows(min_row=start_row):
        cell = row[header_col_index - 1]
        if cell.value:
            header_key = str(cell.value).strip()
            if header_key in header_style_map:
                template_row = header_style_map[header_key]
                copy_style_from_template_row(template_row, row)

    output_wb.save(output_path)

